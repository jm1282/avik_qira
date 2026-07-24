"""Fill + validate the QA workbook and assemble a local Test Run folder.

Per scenario it:
  * waits (smartly) for the traceability folder's files to finish (Log + SSML +
    Audio can take 2-5 min): SSML must be valid `<speak>` XML, the Audio file
    size must be non-zero and stable across polls, and the Log must exist;
  * derives Output Text from the SSML `<lang>` text (folder file preferred, else
    the logcat-captured SSML in the manifest);
  * validates INPUT (exact prompt incl. special chars, on-device + vs Input Text
    sheet) and OUTPUT (SSML output text vs on-screen "screenshot" response text);
  * writes Output Text / Status / File Name / Device Info / Audio File Name and a
    validation note into the results workbook (only existing columns are used);
  * copies the artifacts into a local Test Run folder with proper file names
    (Phone_<code>_<usecase>.{ssml,wav,log,txt}) for audio validation.

Usage:
    python vv_fill_results.py --xlsx "<results.xlsx>" \
        --input-xlsx "<Input Text 1.xlsx>" [--tabs zh_CN] \
        [--test-run-dir "Test Run"] [--file-wait 300]
"""

import argparse
import difflib
import json
import os
import re
import shutil
import subprocess
import sys
import time
import unicodedata
import xml.etree.ElementTree as ET

import openpyxl

import vv_config as cfg

AUDIO_EXTS = (".wav", ".mp3", ".m4a", ".ogg", ".pcm", ".aac", ".flac")
SSML_EXTS = (".ssml", ".xml")
LOG_EXTS = (".log", ".txt")
EXCEL_MAX = 32000
OUTPUT_MATCH_THRESHOLD = 0.60


def adb_base(serial):
    return ["adb"] + (["-s", serial] if serial else [])


def adb_pull(adb, src, dst):
    os.makedirs(os.path.dirname(dst) or ".", exist_ok=True)
    return subprocess.run(adb + ["pull", src, dst], capture_output=True, text=True)


def adb_ls(adb, path):
    r = subprocess.run(adb + ["shell", "ls", "-1", path], capture_output=True, text=True)
    out = []
    for line in (r.stdout or "").splitlines():
        s = line.strip()
        if s and not s.startswith("ls:"):
            out.append(s)
    return out


# ---- text helpers --------------------------------------------------------

def norm_ws(s):
    return re.sub(r"[ \t]+", " ", re.sub(r"\s*\n\s*", "\n", (s or "").strip()))


def normalize_for_match(s):
    if not s:
        return ""
    s = s.replace("\u200e", "").replace("\u200f", "")
    return "".join(ch.lower() for ch in s if unicodedata.category(ch)[0] in ("L", "N"))


def match_ratio(a, b):
    na, nb = normalize_for_match(a), normalize_for_match(b)
    if not na or not nb:
        return 0.0
    if na in nb or nb in na:
        return 1.0
    return difflib.SequenceMatcher(None, na, nb).ratio()


def ssml_to_text(raw):
    if not raw:
        return ""
    try:
        return norm_ws("".join(ET.fromstring(raw).itertext()))
    except Exception:
        return norm_ws(re.sub(r"<[^>]+>", " ", raw))


def read_text(path):
    if not path or not os.path.isfile(path):
        return ""
    with open(path, "rb") as f:
        return f.read().decode("utf-8", errors="replace")


def sanitize(name):
    return re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", (name or "").strip()).rstrip(". ")


def classify_files(folder):
    ssml = audio = log = None
    if not os.path.isdir(folder):
        return ssml, audio, log
    for name in sorted(os.listdir(folder)):
        low = name.lower()
        full = os.path.join(folder, name)
        if not os.path.isfile(full):
            continue
        if ssml is None and (low.endswith(SSML_EXTS) or "ssml" in low):
            ssml = full
        elif audio is None and low.endswith(AUDIO_EXTS):
            audio = full
        elif log is None and (low.endswith(LOG_EXTS) or "log" in low):
            log = full
    return ssml, audio, log


# ---- folder resolution + smart completeness wait -------------------------

def poll_new_folder(adb, trace_dir, seen, timeout_s):
    deadline = time.time() + timeout_s
    while time.time() < deadline:
        for name in adb_ls(adb, '"%s"' % trace_dir):
            if name not in seen:
                seen.add(name)
                return trace_dir.rstrip("/") + "/" + name
        time.sleep(5)
    return None


def wait_folder_complete(adb, folder_dev, local_parent, timeout_s):
    """Pull + poll until SSML valid, Audio non-zero & stable, Log present."""
    local = os.path.join(local_parent, os.path.basename(folder_dev.rstrip("/")))
    deadline = time.time() + timeout_s
    prev_audio = -1
    complete = False
    while True:
        adb_pull(adb, folder_dev, local_parent)
        ssml_f, audio_f, log_f = classify_files(local)
        ssml_ok = bool(ssml_f) and "<speak" in read_text(ssml_f).lower()
        audio_size = os.path.getsize(audio_f) if audio_f and os.path.isfile(audio_f) else 0
        audio_ok = audio_size > 0 and audio_size == prev_audio
        if ssml_ok and audio_ok and log_f:
            complete = True
            break
        prev_audio = audio_size
        if time.time() >= deadline:
            break
        time.sleep(min(10, max(2, int(timeout_s / 20))))
    return local, complete


# ---- workbook helpers ----------------------------------------------------

def find_col(ws, header_candidates):
    wanted = [c.strip().lower() for c in header_candidates]
    for col in range(1, ws.max_column + 1):
        val = ("" if ws.cell(cfg.HEADER_ROW, col).value is None
               else str(ws.cell(cfg.HEADER_ROW, col).value).strip().lower())
        if val and val in wanted:
            return col
    return None


def ensure_col(ws, header):
    col = find_col(ws, [header])
    if col is not None:
        return col
    col = ws.max_column + 1
    ws.cell(row=cfg.HEADER_ROW, column=col, value=header)
    return col


def clip(s):
    s = s or ""
    return s if len(s) <= EXCEL_MAX else s[:EXCEL_MAX] + "\n...[truncated]"


# ---- per-tab fill + validate + assemble ----------------------------------

def fill_tab(adb, wb, tab, meta, work_dir, input_texts, test_run_root,
             new_folder_wait, file_wait):
    manifest_dev = "%s/%s_manifest.json" % (cfg.DEVICE_OUTPUT_DIR, tab)
    manifest_local = os.path.join(work_dir, "%s_manifest.json" % tab)
    adb_pull(adb, manifest_dev, manifest_local)
    if not os.path.isfile(manifest_local):
        print("  [%s] no manifest on device. Skipping." % tab)
        return 0
    with open(manifest_local, "r", encoding="utf-8") as f:
        manifest = json.load(f)

    trace_dir = manifest.get("traceLocaleDir", "")
    device_info = manifest.get("deviceInfo", "")
    seen = set(adb_ls(adb, '"%s"' % trace_dir)) if trace_dir else set()

    ws = wb[tab]
    col_out = ensure_col(ws, meta["output_text_header"])
    col_ssml = find_col(ws, [meta["output_ssml_header"]])   # write only if present
    col_audio = find_col(ws, cfg.AUDIO_FILE_HEADERS)
    col_status = find_col(ws, cfg.STATUS_HEADERS)
    col_devinfo = find_col(ws, cfg.DEVICE_INFO_HEADERS)
    col_file = find_col(ws, cfg.FILE_NAME_HEADERS)
    col_detail = find_col(ws, cfg.DETAIL_HEADERS)
    tab_inputs = input_texts.get(tab, {})
    tab_run_dir = os.path.join(test_run_root, tab)
    os.makedirs(tab_run_dir, exist_ok=True)

    filled = 0
    for res in manifest.get("results", []):
        row = res.get("row", -1)
        if row is None or row < cfg.FIRST_DATA_ROW:
            continue
        use_case = res.get("useCase", "")
        input_text = res.get("inputText", "")
        response_text = res.get("responseText", "")
        base = cfg.build_file_name(tab, row, use_case)

        # --- resolve trace folder + wait for completeness ---
        folder_dev = res.get("traceFolder", "")
        if not folder_dev and trace_dir and new_folder_wait > 0:
            folder_dev = poll_new_folder(adb, trace_dir, seen, new_folder_wait) or ""
        ssml_raw = ""
        audio_local = log_local = None
        files_complete = None
        if folder_dev:
            local, files_complete = wait_folder_complete(
                adb, folder_dev, os.path.join(work_dir, tab), file_wait)
            ssml_f, audio_local, log_local = classify_files(local)
            ssml_raw = read_text(ssml_f)

        if not ssml_raw:
            ssml_raw = res.get("ssml", "") or ""
        out_text = ssml_to_text(ssml_raw) or res.get("ssmlText", "") or norm_ws(response_text)

        # --- validation ---
        sheet_input = tab_inputs.get(row, "")
        input_ok = bool(res.get("inputVerified", False))
        if sheet_input:
            input_ok = input_ok and (
                normalize_for_match(input_text) == normalize_for_match(sheet_input))
        out_ratio = match_ratio(out_text, response_text)
        out_ok = out_ratio >= OUTPUT_MATCH_THRESHOLD
        has_output = bool(out_text)
        status = ("Pass" if (has_output and input_ok and out_ok)
                  else (res.get("status", "NoOutput") if not has_output else "Review"))

        # --- assemble Test Run artifacts (proper file names) ---
        audio_name = ""
        if ssml_raw:
            with open(os.path.join(tab_run_dir, base + ".ssml"), "w", encoding="utf-8") as f:
                f.write(ssml_raw)
        if audio_local and os.path.isfile(audio_local):
            ext = os.path.splitext(audio_local)[1] or ".wav"
            audio_name = base + ext
            shutil.copy2(audio_local, os.path.join(tab_run_dir, audio_name))
        if log_local and os.path.isfile(log_local):
            shutil.copy2(log_local, os.path.join(tab_run_dir, base + ".log"))
        with open(os.path.join(tab_run_dir, base + ".txt"), "w", encoding="utf-8") as f:
            f.write("Use case: %s\nFile Name: %s\nStatus: %s\n"
                    "Input Match: %s\nOutput vs Screen Match: %s (%d%%)\n\n"
                    "Input Text:\n%s\n\nOutput Text (SSML):\n%s\n\nOn-screen Response:\n%s\n"
                    % (use_case, base, status, "Yes" if input_ok else "No",
                       "Yes" if out_ok else "No", round(out_ratio * 100),
                       input_text, out_text, response_text))

        # --- write workbook (existing columns only) ---
        ws.cell(row=row, column=col_out, value=clip(out_text))
        if col_ssml and ssml_raw:
            ws.cell(row=row, column=col_ssml, value=clip(ssml_raw))
        if col_audio and audio_name:
            ws.cell(row=row, column=col_audio, value=audio_name)
        if col_file:
            ws.cell(row=row, column=col_file, value=base)
        if col_status:
            ws.cell(row=row, column=col_status, value=status)
        if col_devinfo and device_info and not str(
                ws.cell(row=row, column=col_devinfo).value or "").strip():
            ws.cell(row=row, column=col_devinfo, value=device_info)
        if col_detail:
            note = "VV: Input Match=%s | Output vs Screen=%s (%d%%) | Status=%s" % (
                "Yes" if input_ok else "No", "Yes" if out_ok else "No",
                round(out_ratio * 100), status)
            existing = str(ws.cell(row=row, column=col_detail).value or "").strip()
            ws.cell(row=row, column=col_detail,
                    value=(existing + " | " + note) if existing else note)

        filled += 1
        print("  [%s] row %d %-16s status=%s input=%s out=%d%% files=%s"
              % (tab, row, base, status, "Y" if input_ok else "N", round(out_ratio * 100),
                 ("complete" if files_complete else "none" if files_complete is None
                  else "incomplete")))
    return filled


def main():
    ap = argparse.ArgumentParser(
        description="Fill + validate QA workbook and build a Test Run folder.")
    ap.add_argument("--xlsx", required=True, help="Results workbook (modified in place).")
    ap.add_argument("--input-xlsx", default="", help="Input Text workbook (for input validation).")
    ap.add_argument("--tabs", default="", help="Comma-separated tabs (default: all configured).")
    ap.add_argument("--work-dir", default="run", help="Local dir for pulled artifacts.")
    ap.add_argument("--test-run-dir", default="Test Run", help="Root for the Test Run output.")
    ap.add_argument("--output", default="", help="Write workbook to this path instead of in place.")
    ap.add_argument("--file-wait", type=int, default=300,
                    help="Max seconds to wait for a bound folder's files to complete.")
    ap.add_argument("--new-folder-wait", type=int, default=30,
                    help="Max seconds to wait for an unbound folder to appear.")
    ap.add_argument("--serial", default="", help="adb device serial (optional).")
    args = ap.parse_args()

    tabs = [t.strip() for t in args.tabs.split(",") if t.strip()] or list(cfg.locales().keys())
    adb = adb_base(args.serial)
    os.makedirs(args.work_dir, exist_ok=True)
    input_texts = cfg.read_input_texts(args.input_xlsx) if args.input_xlsx else {}

    stamp = time.strftime("%Y-%m-%d_%H-%M-%S")
    test_run_root = os.path.join(args.test_run_dir, stamp)
    os.makedirs(test_run_root, exist_ok=True)

    if not args.output:
        backup = "%s.bak-%s" % (args.xlsx, stamp)
        shutil.copy2(args.xlsx, backup)
        print("Backup written: %s" % backup)

    wb = openpyxl.load_workbook(args.xlsx)
    total = 0
    for tab in tabs:
        meta = cfg.locales().get(tab)
        if not meta or tab not in wb.sheetnames:
            print("Skipping tab (unknown or missing): %s" % tab)
            continue
        total += fill_tab(adb, wb, tab, meta, args.work_dir, input_texts, test_run_root,
                          args.new_folder_wait, args.file_wait)

    dest = args.output or args.xlsx
    wb.save(dest)
    # Copy the filled workbook into the Test Run folder too.
    try:
        shutil.copy2(dest, os.path.join(test_run_root, os.path.basename(dest)))
    except Exception as e:
        print("WARN: could not copy workbook into Test Run: %s" % e)
    print("Saved %s (%d rows)" % (dest, total))
    print("Test Run folder: %s" % os.path.abspath(test_run_root))


if __name__ == "__main__":
    sys.exit(main())
