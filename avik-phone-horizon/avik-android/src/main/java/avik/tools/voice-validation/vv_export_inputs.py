"""Export per-locale Voice Validation input JSON from the Input Text workbook.

Reads the dedicated "Voice Validation Input Text 1.xlsx" (single "Instruction"
sheet, one BCP-47 column per locale), extracts the localized prompt for every
scenario, and writes ``<out_dir>/<tab>.json`` in the schema the on-device
``VoiceValidationInput`` reads. Row numbers are the *results* workbook rows
(input row + RESULT_ROW_OFFSET) so the manifest maps straight onto the QA sheet.

Usage:
    python vv_export_inputs.py --input-xlsx "<Input Text 1.xlsx>" [--tabs zh_CN] \
        [--out-dir out] [--push]
"""

import argparse
import json
import os
import subprocess
import sys

import openpyxl

import vv_config as cfg


def _norm(v):
    return "" if v is None else str(v).strip()


def _find_col(ws, header_row, header_candidates):
    wanted = [c.strip().lower() for c in header_candidates]
    for col in range(1, ws.max_column + 1):
        val = _norm(ws.cell(row=header_row, column=col).value).lower()
        if val and val in wanted:
            return col
    return None


def _find_locale_col(ws, header_row, locale_tag):
    """Column whose header matches the BCP-47 locale (case-insensitive, - or _)."""
    want = locale_tag.strip().lower().replace("_", "-")
    for col in range(1, ws.max_column + 1):
        val = _norm(ws.cell(row=header_row, column=col).value).lower().replace("_", "-")
        if val == want:
            return col
    return None


def export_tab(ws, tab, meta):
    hr = cfg.INPUT_HEADER_ROW
    locale_col = _find_locale_col(ws, hr, meta["device_locale"])
    if locale_col is None:
        raise SystemExit("Input sheet has no column for locale '%s' (tab %s). Headers: %s"
                         % (meta["device_locale"], tab,
                            [ws.cell(hr, c).value for c in range(1, ws.max_column + 1)]))
    use_case_col = _find_col(ws, hr, cfg.INPUT_USE_CASE_HEADERS)
    criteria_col = _find_col(ws, hr, cfg.INPUT_CRITERIA_HEADERS)

    scenarios = []
    for in_row in range(cfg.INPUT_FIRST_ROW, cfg.INPUT_LAST_ROW + 1):
        input_text = _norm(ws.cell(row=in_row, column=locale_col).value)
        if not input_text:
            continue
        scenarios.append({
            "row": in_row + cfg.RESULT_ROW_OFFSET,
            "useCase": _norm(ws.cell(in_row, use_case_col).value) if use_case_col else "",
            "inputText": input_text,
            "criteria": _norm(ws.cell(in_row, criteria_col).value) if criteria_col else "",
        })

    return {
        "tab": tab,
        "responseLanguage": meta["response_language"],
        "appLocale": meta["device_locale"],
        "voiceLocale": meta["voice_locale"],
        "affirmative": meta["affirmative"],
        "voiceModelDir": cfg.VOICE_MODEL_DIR,
        "traceBase": cfg.TRACE_BASE,
        "languageCandidates": meta["language_candidates"],
        "scenarios": scenarios,
    }


def main():
    ap = argparse.ArgumentParser(description="Export Voice Validation input JSON.")
    ap.add_argument("--input-xlsx", required=True, help="Path to the Input Text workbook.")
    ap.add_argument("--tabs", default="", help="Comma-separated tabs (default: all configured).")
    ap.add_argument("--out-dir", default="out", help="Local output directory.")
    ap.add_argument("--push", action="store_true", help="adb push the JSON to the device.")
    ap.add_argument("--serial", default="", help="adb device serial (optional).")
    args = ap.parse_args()

    tabs = [t.strip() for t in args.tabs.split(",") if t.strip()] or list(cfg.locales().keys())
    os.makedirs(args.out_dir, exist_ok=True)
    wb = openpyxl.load_workbook(args.input_xlsx, data_only=True)
    if cfg.INPUT_SHEET not in wb.sheetnames:
        raise SystemExit("Input workbook has no '%s' sheet. Sheets: %s"
                         % (cfg.INPUT_SHEET, wb.sheetnames))
    ws = wb[cfg.INPUT_SHEET]

    adb = ["adb"] + (["-s", args.serial] if args.serial else [])
    if args.push:
        subprocess.run(adb + ["shell", "mkdir", "-p", cfg.DEVICE_INPUT_DIR], check=False)

    for tab in tabs:
        meta = cfg.locales().get(tab)
        if not meta:
            print("Skipping unknown tab: %s" % tab)
            continue
        data = export_tab(ws, tab, meta)
        out_path = os.path.join(args.out_dir, "%s.json" % tab)
        with open(out_path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        print("Wrote %s (%d scenarios)" % (out_path, len(data["scenarios"])))
        if args.push:
            dest = "%s/%s.json" % (cfg.DEVICE_INPUT_DIR, tab)
            r = subprocess.run(adb + ["push", out_path, dest], capture_output=True, text=True)
            print("  push -> %s : %s" % (dest, (r.stdout or r.stderr).strip()))


if __name__ == "__main__":
    sys.exit(main())
