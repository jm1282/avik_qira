"""Shared configuration for the Qira Voice Validation host harness.

Single source of truth for the per-tab locale contract consumed by both the
exporter (device input JSON) and the results filler (Excel columns). Keeping it
here avoids locale-specific string tables leaking into the instrumentation.
"""

# On-device paths (must match the Java side defaults).
DEVICE_INPUT_DIR = "/sdcard/avik/voicevalidation/input"
DEVICE_OUTPUT_DIR = "/sdcard/avik/voicevalidation/output"
TRACE_BASE = "/sdcard/Download/Voice Traceability"
VOICE_MODEL_DIR = "en-US-Ava-preview_DragonHDLatestNeural"

# The QA (results) workbook header row and first/last data rows (1-based).
HEADER_ROW = 2
FIRST_DATA_ROW = 3
LAST_DATA_ROW = 14

# The dedicated Input Text workbook ("Voice Validation Input Text 1.xlsx"):
# a single "Instruction" sheet with one BCP-47 column per locale.
#   row 1  = headers (Use case | Text to input (english) | Evaluate | <locale cols...>)
#   rows 2..13 = the 12 scenarios (same order as the results tabs)
# A results-file data row = input row + RESULT_ROW_OFFSET.
INPUT_SHEET = "Instruction"
INPUT_HEADER_ROW = 1
INPUT_FIRST_ROW = 2
INPUT_LAST_ROW = 13
RESULT_ROW_OFFSET = 1
INPUT_USE_CASE_HEADERS = ["Use case", "Use Case"]
INPUT_CRITERIA_HEADERS = ["Evaluate", "Evaluate criteria"]

# Per workbook tab.
#   device_locale      : BCP-47 tag applied to the device system locale (reboot).
#   voice_locale       : traceability sub-folder name under the voice model dir.
#   response_language  : English label of Qira's response-language setting.
#   affirmative        : reply sent when a response asks a counter-question.
#   language_candidates: labels/keywords used to find/verify the language row
#                        across an English or localized Qira UI.
#   input_headers      : accepted header texts for the localized prompt column.
#   output_text_header : header used for the SSML-derived output text.
#   output_ssml_header : header used for the raw SSML.
LOCALES = {
    "zh_CN": {
        "device_locale": "zh-CN",
        "voice_locale": "zh-CN",
        "response_language": "Chinese (China)",
        "affirmative": "yes",
        "language_candidates": [
            "Chinese (China)", "\u4e2d\u6587\uff08\u4e2d\u56fd\uff09",
            "\u4e2d\u6587", "\u4e2d\u56fd", "\u7b80\u4f53\u4e2d\u6587",
        ],
        "input_headers": ["Text to input (Chinese)"],
        "output_text_header": "Output Text (Chinese)",
        "output_ssml_header": "Output SSML",
    },
    "ja_JP": {
        "device_locale": "ja-JP",
        "voice_locale": "ja-JP",
        "response_language": "Japanese (Japan)",
        "affirmative": "yes",
        "language_candidates": [
            "Japanese (Japan)", "\u65e5\u672c\u8a9e\uff08\u65e5\u672c\uff09",
            "\u65e5\u672c\u8a9e", "\u65e5\u672c",
        ],
        "input_headers": ["Text to input (Japanese)"],
        "output_text_header": "Output Text (Japanese)",
        "output_ssml_header": "Output SSML",
    },
    "ro_RO": {
        "device_locale": "ro-RO",
        "voice_locale": "ro-RO",
        "response_language": "Romanian (Romania)",
        "affirmative": "yes",
        "language_candidates": [
            "Romanian (Romania)", "Rom\u00e2n\u0103 (Rom\u00e2nia)",
            "Rom\u00e2n\u0103", "rom\u00e2n\u0103", "Romanian",
        ],
        "input_headers": ["Text to input (Romania)", "Text to input (Romanian)"],
        "output_text_header": "Output Text (Romania)",
        "output_ssml_header": "Output SSML",
    },
    "pl_PL": {
        "device_locale": "pl-PL",
        "voice_locale": "pl-PL",
        "response_language": "Polish (Poland)",
        "affirmative": "yes",
        "language_candidates": [
            "Polish (Poland)", "Polski (Polska)", "Polski", "polski", "Polish",
        ],
        "input_headers": ["Text to input (Polish)"],
        "output_text_header": "Output Text (Polish)",
        "output_ssml_header": "Output SSML",
    },
    "ar_SA": {
        "device_locale": "ar-SA",
        "voice_locale": "ar-SA",
        "response_language": "Arabic (Saudi Arabia)",
        "affirmative": "yes",
        "language_candidates": [
            "Arabic (Saudi Arabia)",
            "\u0627\u0644\u0639\u0631\u0628\u064a\u0629 "
            "(\u0627\u0644\u0645\u0645\u0644\u0643\u0629 \u0627\u0644\u0639\u0631\u0628\u064a\u0629 "
            "\u0627\u0644\u0633\u0639\u0648\u062f\u064a\u0629)",
            "\u0627\u0644\u0639\u0631\u0628\u064a\u0629", "Arabic",
        ],
        "input_headers": ["Text to input (Arabic)"],
        "output_text_header": "Output Text (Arabic)",
        "output_ssml_header": "Output SSML",
    },
}

# Headers used to locate common columns (matched case-insensitively, trimmed).
USE_CASE_HEADERS = ["Use case", "Use Case"]
CRITERIA_HEADERS = ["Evaluate criteria", "Evaluate"]
STATUS_HEADERS = ["Status"]
FILE_NAME_HEADERS = ["File Name"]
DEVICE_INFO_HEADERS = ["Device Info"]
AUDIO_FILE_HEADERS = ["Audio File Name"]

# Column where the filler records the input/output validation note.
DETAIL_HEADERS = ["Detail Observation/Validation", "Comment"]

# Artifact / File Name convention (matches the existing sheet: Phone_CH_Weather).
LOCALE_CODE = {"zh_CN": "CH", "ja_JP": "JP", "ro_RO": "RO", "pl_PL": "PL", "ar_SA": "Ar"}
ROW_SHORT_NAME = {
    3: "Weather", 4: "CMU", 5: "Recomendations", 6: "Ask Anything",
    7: "Math1", 8: "Math2", 9: "Complex vocabulary", 10: "Poetry",
    11: "Support", 12: "Encouragement", 13: "Sadness", 14: "Any Input",
}


def build_file_name(tab, row, use_case=""):
    """Base name for a scenario's artifacts, e.g. Phone_CH_Weather."""
    code = LOCALE_CODE.get(tab, tab)
    short = ROW_SHORT_NAME.get(row) or (use_case.strip() if use_case else "row%d" % row)
    return "Phone_%s_%s" % (code, short)


def locales():
    return LOCALES


def read_input_texts(input_xlsx_path):
    """{tab: {results_row: exact_input_text}} read from the Input Text workbook."""
    import openpyxl

    def norm(v):
        return "" if v is None else str(v).strip()

    wb = openpyxl.load_workbook(input_xlsx_path, data_only=True)
    if INPUT_SHEET not in wb.sheetnames:
        return {}
    ws = wb[INPUT_SHEET]
    out = {}
    for tab, meta in LOCALES.items():
        want = meta["device_locale"].lower().replace("_", "-")
        col = None
        for c in range(1, ws.max_column + 1):
            if norm(ws.cell(INPUT_HEADER_ROW, c).value).lower().replace("_", "-") == want:
                col = c
                break
        if col is None:
            continue
        rowmap = {}
        for in_row in range(INPUT_FIRST_ROW, INPUT_LAST_ROW + 1):
            t = norm(ws.cell(in_row, col).value)
            if t:
                rowmap[in_row + RESULT_ROW_OFFSET] = t
        out[tab] = rowmap
    return out
